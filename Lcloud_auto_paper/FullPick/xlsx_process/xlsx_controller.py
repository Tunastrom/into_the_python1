import openpyxl
from FullPick.common import percent_printer

def xlsx_controller(one_time_fullbackup_dict, xlsx_path_dir):
    xlsx_name = 'Lcloud_backup_excel_202105'
    xlsx_file = openpyxl.load_workbook(xlsx_path_dir+xlsx_name+'.xlsx', read_only=False, data_only=True)
    sheet_2 = xlsx_file["클라이언트별 1회 Full 백업량"]
    max_row, switch, checker = sheet_2.max_row+1, bool, 'full_change'
    sitename_set = set()
    # 엑셀파일의 1행씩 움직이며 one_time_fullbackup_dict의 전체요소와 비교
    print("<full 백업용량 수정 시작>")
    switch = False
    for row_index in range(1,max_row):
        sitename_cell, servername_cell, backuptype_cell, backupdata_cell, checking_cell = sheet_2[f'A{row_index}'], sheet_2[f'B{row_index}'], sheet_2[f'C{row_index}'], sheet_2.cell(row=row_index, column=4), sheet_2[f'E{row_index}']
        # print(servername_cell.value, backuptype_cell.value, backupdata_cell.value)
        for policy_name,backupdata in one_time_fullbackup_dict.items():
            # 엑셀파일의 '서버명' 컬럼 값과 one_time_fullbackup_dict의 정책명이 일치하는 결과가 있을 때
            if servername_cell.value is not None and policy_name.find(servername_cell.value) != -1:
                # 백업용량 수정한 사이트 명 추가
                sitename_set.add(sitename_cell.value)
                # file 백업 정책으로 filesystem 백업받은 경우
                if policy_name.lower().find('_fs') != -1 and backuptype_cell.value.lower() == 'filesystem':
                    backupdata_cell.value = int(round(backupdata['Gigabyte'],0))
                    checking_cell.value = 'changed'
                # file 백업 정책으로 application 또는 NAS 백업받은 경우
                elif policy_name.lower().find('_fs') != -1 and backuptype_cell.value.lower() != 'filesystem':
                    backupdata_cell.value = int(round(backupdata['Gigabyte'],0))
                    checking_cell.value = 'changed'
                # file 백업 정책이 아닌 것으로 application(Oracle, MSSQL) 백업받은 경우
                elif policy_name.lower().find('_fs') == -1 and backuptype_cell.value.lower() != 'filesystem':
                    backupdata_cell.value = int(round(backupdata['Gigabyte'],0))
                    checking_cell.value = 'changed'
                # print(servername_cell.value, backuptype_cell.value, backupdata_cell.value)
        # 진행률 표시
        percent = round((row_index/max_row)*100)
        switch = percent_printer(percent,switch, checker)

#    print("<backup기록 없는 서버 삭제 시작>")
#    switch, checker = False, 'delete_row'
#    for row_index in range(1, max_row):
#        sitename_cell = sheet_2[f'A{row_index}']
#        checking_cell = sheet_2[f'E{row_index}']
#        # 엑셀파일의 '서버명' 컬럼 값과 one_time_fullbackup_dict의 정책명이 일치하는 결과가 없을 때 해당 row 제거
#        if sitename_cell.value in sitename_set and checking_cell.value is None:
#            sheet_2.delete_rows(row_index)
#        elif sitename_cell.value not in sitename_set:
#            sitename_set.add(sitename_cell.value)
#        percent = round((row_index/max_row)*100)
#        switch = percent_printer(percent,switch, checker)
    #     elif sitename_cell.value in sitename_set and checking_cell.value == 'changed':
    #         checking_cell.value = None


    sitename_dict = dict()

    for sitename in sitename_set:
        sitename_dict.setdefault(sitename,0)

    for row_index in range(2, max_row):
        sitename_cell = sheet_2[f'A{row_index}']
        backupdata_cell = sheet_2.cell(row=row_index, column=4)
        if sitename_cell.value in sitename_dict.keys() and backupdata_cell.value is not None:
            sitename_dict[sitename_cell.value] += backupdata_cell.value

    xlsx_file.save(xlsx_path_dir+f'Lcloud_backup_excel_202105.xlsx')





